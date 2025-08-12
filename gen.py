import os
import shutil

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter


class imageSet:
    def __init__(self, path, mode, xlsx_path):
        self.path = path
        self.xlsx = xlsx_path
        self.mode = mode
        self.pics = os.listdir(self.path)  # 排序
        self.pics.sort(key=lambda x: int(x.split('.')[0]))

    def run(self):
        # 打开现有的Excel工作簿
        # wb = load_workbook("/Users/ericmac/qrcode/8.21/妇幼/妇幼家具修改1.xlsx")
        wb = load_workbook(self.xlsx)
        # 获取默认的工作表
        ws = wb.active
        shift = 2
        df = pd.read_csv('temp.csv')
        titles_len = len(list(df.head(0)))
        os.remove('temp.csv')
        for row in range(shift, shift + len(self.pics)):
            ws.row_dimensions[row].height = 150
            col_letter = get_column_letter(titles_len+1)
            ws.column_dimensions[col_letter].width = 40
            ins = Image(f'{self.path}/' + self.pics[row - shift])
            ins.height, ins.width = 150, 150
            anc = col_letter
            # print(ins)
            ws.add_image(ins, anchor=anc + str(row))

        # wb.save('/Users/ericmac/qrcode/8.21/妇幼/妇幼家具修改1.xlsx')
        wb.save(self.xlsx)
        if self.mode == 1:
            shutil.rmtree(self.path)
#
# path = r"output/妇幼修改家具"
# pics = os.listdir(path)
# pics.sort(key=lambda x: int(x.split('.')[0]))
# print(pics)

#
# # 打开现有的Excel工作簿
# wb = load_workbook("/Users/ericmac/qrcode/8.21/妇幼/妇幼家具修改1.xlsx")
# # 获取默认的工作表
# ws = wb.active
# shift = 2
# for row in range(shift, shift+len(pics)):
#     ws.row_dimensions[row].height = 150
#     col_letter = get_column_letter(8)
#     ws.column_dimensions[col_letter].width = 40
#     ins = Image(f'{path}/' + pics[row-shift])
#     ins.height, ins.width = 150, 150
#     ws.add_image(ins, anchor='I'+str(row))
#
# wb.save('/Users/ericmac/qrcode/8.21/妇幼/妇幼家具修改1.xlsx')
